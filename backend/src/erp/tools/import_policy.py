"""禁售政策导入 CLI（部署机 api 容器内执行，读本地文件灌入 refdata.prohibited_policy）。

用法：
  docker compose -f infra/docker-compose.yml exec api \\
    python -m erp.tools.import_policy --file /data/prohibited_policy.csv

数据源 = 飞书「沃尔玛禁止」表 OJSrkV 导出。列名任一即可（含常见别名）：
  category_en        ← category_en | category            （幂等键，缺失该行计 err）
  seq                ← seq | id                           （L3 prompt 排序，源仓 id）
  category_zh        ← category_zh | category_cn
  overall_status     ← overall_status | status
  prohibited_items   ← prohibited_items | prohibited      （L3 prompt 摘要取此列）
  conditional_items / preapproval_items / zh_seller_risk / zh_seller_notes

政策为全局参考数据（refdata，无团队），走超管 system_tx。幂等：以 category_en 为键
ON CONFLICT 更新，重复导入不产生新行（计 ok）。导入后 L3 system prompt 自动拼入 37 条
静态政策块（吃 provider prefix cache）。
"""

import argparse
import asyncio
import csv
import json
import sys
from pathlib import Path
from typing import Any

from erp.compliance import import_service
from erp.core.db import get_session_factory, system_tx


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".jsonl":
        return [json.loads(line) for line in path.read_text().splitlines() if line.strip()]
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            return [dict(r) for r in csv.DictReader(f)]
    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import (  # type: ignore[import-untyped]  # noqa: PLC0415 可选依赖延迟导入
                load_workbook,
            )
        except ImportError:
            print("读取 xlsx 需要 openpyxl；请改用 csv/jsonl，或容器内 pip install openpyxl")
            raise
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        out: list[dict[str, Any]] = []
        for r in rows_iter:
            out.append({header[i]: r[i] for i in range(len(header)) if header[i]})
        return out
    raise ValueError(f"不支持的文件格式：{suffix}（用 csv/xlsx/jsonl）")


async def _run(path: Path) -> dict[str, Any]:
    rows = _read_rows(path)
    if not rows:
        print("文件无数据行")
        return {"ok": 0, "skip": 0, "err": 0}
    sessions = get_session_factory()
    async with system_tx(sessions) as s:
        job = await import_service.create_job(
            s,
            domain=import_service.POLICY_DOMAIN,
            source_name=path.name,
            total_rows=len(rows),
            fmt="jsonl" if path.suffix.lower() == ".jsonl" else path.suffix.lower().lstrip("."),
        )
    job_id = job["id"]
    try:
        async with system_tx(sessions) as s:
            summary = await import_service.import_rows(s, job_id=job_id, rows=rows)
    except Exception as e:
        async with system_tx(sessions) as s:
            await import_service.mark_failed(s, job_id=job_id, error=str(e))
        print(f"导入失败（job {job_id}）：{e}")
        raise
    print(
        f"导入完成 job {job_id} [policy]："
        f"新增/更新 {summary['ok']} / 跳过 {summary['skip']} / 错误 {summary['err']}"
        f"（共 {summary['total']} 行）"
    )
    return summary


def main() -> int:
    p = argparse.ArgumentParser(description="ERP 禁售政策导入（refdata.prohibited_policy）")
    p.add_argument("--file", required=True, help="本地文件（csv/xlsx/jsonl）")
    args = p.parse_args()
    path = Path(args.file)
    if not path.is_file():
        print(f"文件不存在：{path}", file=sys.stderr)
        return 2
    try:
        asyncio.run(_run(path))
    except Exception:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
