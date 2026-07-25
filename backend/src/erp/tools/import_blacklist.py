"""黑名单导入 CLI（R2-02；部署机 api 容器内执行，读本地文件灌入 app.blacklist_*）。

用法（api 容器无 volume 挂载，先 cp 进 /tmp 再 exec；运维全流程见
infra/local-deploy/README.md「黑名单 / TRO bulk 导入」）：
  docker compose -f infra/docker-compose.yml cp ./brands.csv api:/tmp/
  docker compose -f infra/docker-compose.yml exec api \\
    python -m erp.tools.import_blacklist --domain blacklist_brand --file /tmp/brands.csv

文件格式（按扩展名识别 csv/xlsx/jsonl）；列名任一即可（大小写敏感按下表）：
  blacklist_brand    → brand[, brand_display, reason]
  blacklist_seller   → seller_id[, seller_name, reason]
  blacklist_asin     → asin[, reason]
  blacklist_category → category[, reason]
  blacklist_address  → street/address/地址[, reason]（R2-05 钓鱼检；BR-ORD-005 口径归一化）
  blacklist_zip      → zip/zipcode/邮编[, reason]（取前 5 位）
  tro                → case_no[, plaintiff, court, filed_date, law_firm, brand_terms,
                       status, raw_ref]（R2-12 增量2）
                       brand_terms 为 JSON 数组或分号分隔串（词内逗号不拆）；
                       status ∈ active/dismissed/settled（缺省 active）；
                       active 案派生全局 tro_sync 品牌断言，dismissed/settled 撤销该案断言；
                       TRO 拉黑恒为全局，--team 对本域无效。
  注意：lark 钓鱼地址表表头在第 5 行（data-survey SYNTHESIS:44），导出 csv 前先删表头前噪声行。

全局黑名单（team_id=NULL）默认导入；--team <id> 可限定团队。归一化与审核 L0
查表一致，导入即生效。幂等：重复行跳过（skip），占位符品牌（unbranded 等）跳过。
"""

import argparse
import asyncio
import csv
import json
import sys
from pathlib import Path
from typing import Any

from erp.compliance import import_service
from erp.core.db import get_session_factory, system_tx


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".jsonl":
        return [json.loads(line) for line in path.read_text().splitlines() if line.strip()]
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            return [dict(r) for r in csv.DictReader(f)]
    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import (  # type: ignore[import-untyped]  # noqa: PLC0415 可选依赖延迟导入
                load_workbook,
            )
        except ImportError:
            print("读取 xlsx 需要 openpyxl；请改用 csv/jsonl，或容器内 pip install openpyxl")
            raise
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        out: list[dict[str, Any]] = []
        for r in rows_iter:
            out.append({header[i]: r[i] for i in range(len(header)) if header[i]})
        return out
    raise ValueError(f"不支持的文件格式：{suffix}（用 csv/xlsx/jsonl）")


async def _run(domain: str, path: Path, team_id: int | None) -> dict[str, Any]:
    rows = _read_rows(path)
    if not rows:
        print("文件无数据行")
        return {"ok": 0, "skip": 0, "err": 0}
    sessions = get_session_factory()
    async with system_tx(sessions) as s:
        job = await import_service.create_job(
            s,
            domain=domain,
            source_name=path.name,
            total_rows=len(rows),
            fmt="jsonl" if path.suffix.lower() == ".jsonl" else path.suffix.lower().lstrip("."),
            team_id=team_id,
        )
    job_id = job["id"]
    try:
        async with system_tx(sessions) as s:
            summary = await import_service.import_rows(s, job_id=job_id, rows=rows)
    except Exception as e:
        async with system_tx(sessions) as s:
            await import_service.mark_failed(s, job_id=job_id, error=str(e))
        print(f"导入失败（job {job_id}）：{e}")
        raise
    print(
        f"导入完成 job {job_id} [{domain}]："
        f"新增 {summary['ok']} / 跳过 {summary['skip']} / 错误 {summary['err']}"
        f"（共 {summary['total']} 行）"
    )
    return summary


def main() -> int:
    p = argparse.ArgumentParser(description="ERP 黑名单导入")
    p.add_argument(
        "--domain",
        required=True,
        choices=list(import_service.SUPPORTED_DOMAINS),
        help="目标黑名单域",
    )
    p.add_argument("--file", required=True, help="本地文件（csv/xlsx/jsonl）")
    p.add_argument("--team", type=int, default=None, help="限定团队 id（缺省=全局 team_id NULL）")
    args = p.parse_args()
    path = Path(args.file)
    if not path.is_file():
        print(f"文件不存在：{path}", file=sys.stderr)
        return 2
    try:
        asyncio.run(_run(args.domain, path, args.team))
    except Exception:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
