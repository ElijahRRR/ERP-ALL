"""PT 官方 spec 导入 CLI（R3b 认证字段层 + R2-03 上架全量规格，refdata.pt_spec）。

用法：
  docker compose -f infra/docker-compose.yml exec api \\
    python -m erp.tools.import_pt_spec --file /data/pt_spec.jsonl [--chunk-size 200]

数据来源两代：
  R3b（审核子集）：源仓 walmart_audit.walmart_pt_spec 投影（6,942 行，无 fields）。
  R2-03（上架全量）：tools/extract_mp_item_spec.py 无损产出（fields=per-PT 官方 v5
    原始 schema 节点 + '__orderable__' 伪行）。行大（10-100KB），建议 --chunk-size 200。

文件格式（推荐 jsonl，jsonb 列原生保真）；列：
  walmart_product_type（必填）· has_real_cert · real_cert_fields ·
  has_soft_cert · soft_cert_fields · total_fields · required_count · required_fields ·
  fields（可缺省；缺省时不清库中已有值——COALESCE 保留，两代数据可交替重导）

全局参考数据（refdata，无团队属性），走超管 system_tx。幂等：以 walmart_product_type
为键 ON CONFLICT 恒更新，重导不产生新行。
"""

import argparse
import asyncio
import csv
import json
import sys
from pathlib import Path
from typing import Any

from erp.compliance import import_service
from erp.core.db import get_session_factory, system_tx


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".jsonl":
        return [json.loads(line) for line in path.read_text().splitlines() if line.strip()]
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            return [dict(r) for r in csv.DictReader(f)]
    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import (  # type: ignore[import-untyped]  # noqa: PLC0415 可选依赖延迟导入
                load_workbook,
            )
        except ImportError:
            print("读取 xlsx 需要 openpyxl；请改用 csv/jsonl，或容器内 pip install openpyxl")
            raise
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        out: list[dict[str, Any]] = []
        for r in rows_iter:
            out.append({header[i]: r[i] for i in range(len(header)) if header[i]})
        return out
    raise ValueError(f"不支持的文件格式：{suffix}（用 csv/xlsx/jsonl）")


async def _run(path: Path, chunk_size: int = 5000) -> dict[str, Any]:
    rows = _read_rows(path)
    if not rows:
        print("文件无数据行")
        return {"ok": 0, "skip": 0, "err": 0}
    sessions = get_session_factory()
    async with system_tx(sessions) as s:
        job = await import_service.create_job(
            s,
            domain=import_service.PT_SPEC_DOMAIN,
            source_name=path.name,
            total_rows=len(rows),
            fmt="jsonl" if path.suffix.lower() == ".jsonl" else path.suffix.lower().lstrip("."),
            chunk_size=chunk_size,
        )
    job_id = job["id"]
    try:
        async with system_tx(sessions) as s:
            summary = await import_service.import_rows(s, job_id=job_id, rows=rows)
    except Exception as e:
        async with system_tx(sessions) as s:
            await import_service.mark_failed(s, job_id=job_id, error=str(e))
        print(f"导入失败（job {job_id}）：{e}")
        raise
    print(
        f"导入完成 job {job_id} [pt_spec]："
        f"新增/更新 {summary['ok']} / 跳过 {summary['skip']} / 错误 {summary['err']}"
        f"（共 {summary['total']} 行）"
    )
    return summary


def main() -> int:
    p = argparse.ArgumentParser(description="ERP PT spec 导入（refdata.pt_spec，全局参考数据）")
    p.add_argument("--file", required=True, help="本地文件（csv/xlsx/jsonl）")
    p.add_argument(
        "--chunk-size", type=int, default=5000, help="分批提交行数（全量 fields jsonl 建议 200）"
    )
    args = p.parse_args()
    path = Path(args.file)
    if not path.is_file():
        print(f"文件不存在：{path}", file=sys.stderr)
        return 2
    try:
        asyncio.run(_run(path, chunk_size=args.chunk_size))
    except Exception:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
